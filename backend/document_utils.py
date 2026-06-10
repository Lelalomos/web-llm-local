import base64
import io
import json
import os
import subprocess
import tempfile
from pathlib import Path

import openpyxl
import pymupdf
import pytesseract
import requests
import xlrd
from docx import Document
from PIL import Image
from pypdf import PdfReader

from ollama_options import apply_gpu_defaults


SUPPORTED_TEXT_EXTENSIONS = {".txt", ".md", ".json", ".csv"}
SUPPORTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"} | SUPPORTED_TEXT_EXTENSIONS | SUPPORTED_IMAGE_EXTENSIONS
DEFAULT_SUMMARY_PROMPT = "Summarize this document and highlight the key points."
MAX_DOCUMENT_PROMPT_CHARS = 20000
DOCUMENT_CHAT_NUM_CTX = int(os.getenv("DOCUMENT_CHAT_NUM_CTX", "8192"))
PDF_OCR_LANGUAGE = os.getenv("PDF_OCR_LANGUAGE", "eng")
IMAGE_OCR_LANGUAGE = os.getenv("IMAGE_OCR_LANGUAGE", PDF_OCR_LANGUAGE)
PDF_OCR_DPI = int(os.getenv("PDF_OCR_DPI", "200"))
PDF_EXTRACTION_MODE = os.getenv("PDF_EXTRACTION_MODE", "surya_docling")
PDF_ADVANCED_MAX_CHARS = int(os.getenv("PDF_ADVANCED_MAX_CHARS", "50000"))
PDF_SURYA_TIMEOUT_SECONDS = int(os.getenv("PDF_SURYA_TIMEOUT_SECONDS", "60"))
PDF_SURYA_MIN_DOCLING_CHARS = int(os.getenv("PDF_SURYA_MIN_DOCLING_CHARS", "1000"))
PDF_FAST_TEXT_MIN_CHARS = int(os.getenv("PDF_FAST_TEXT_MIN_CHARS", "1000"))
DEFAULT_VISION_OCR_PROMPT = (
    "Extract all visible text from this image. Preserve line breaks and table structure. "
    "Return only the extracted text. If the text is Thai, return Thai text exactly."
)


def extract_document_text(filename: str, content: bytes, app_config: dict | None = None, ollama_url: str | None = None) -> tuple[str, int]:
    cleaned_text, character_count, _metadata = extract_document_text_with_metadata(filename, content, app_config, ollama_url)
    return cleaned_text, character_count


def extract_document_text_with_metadata(filename: str, content: bytes, app_config: dict | None = None, ollama_url: str | None = None) -> tuple[str, int, dict]:
    _, ext = os.path.splitext(filename.lower())
    config = app_config or {}
    metadata = {"ocr_engine_requested": str(config.get("ocr_engine") or "default")}

    if ext not in SUPPORTED_DOCUMENT_EXTENSIONS:
        raise ValueError(f"Unsupported file format '{ext}'")

    if ext == ".pdf":
        extracted_text = _extract_pdf_text(content, filename, config, ollama_url)
        metadata["ocr_engine_used"] = _pdf_extraction_mode(config)
    elif ext == ".docx":
        extracted_text = _extract_docx_text(content)
        metadata["ocr_engine_used"] = "docx"
    elif ext == ".xlsx":
        extracted_text = _extract_xlsx_text(content)
        metadata["ocr_engine_used"] = "xlsx"
    elif ext == ".xls":
        extracted_text = _extract_xls_text(content)
        metadata["ocr_engine_used"] = "xls"
    elif ext in SUPPORTED_IMAGE_EXTENSIONS:
        extracted_text, image_metadata = _extract_image_text_with_metadata(content, config, ollama_url)
        metadata.update(image_metadata)
    else:
        extracted_text = content.decode("utf-8", errors="ignore")
        metadata["ocr_engine_used"] = "text"

    cleaned_text = extracted_text.strip()
    return cleaned_text, len(extracted_text), metadata


def build_document_prompt(file_name: str, file_text: str, user_prompt: str) -> str:
    prompt = (user_prompt or "").strip() or DEFAULT_SUMMARY_PROMPT
    document_body = _truncate_document_text(file_text)
    return (
        f'Context from uploaded file "{file_name}":\n\n'
        "--- START OF FILE CONTENT ---\n"
        f"{document_body}\n"
        "--- END OF FILE CONTENT ---\n\n"
        f"Use the file content above to answer this prompt: {prompt}"
    )


def _extract_pdf_text(content: bytes, filename: str = "document.pdf", app_config: dict | None = None, ollama_url: str | None = None) -> str:
    extraction_mode = _pdf_extraction_mode(app_config or {})
    if extraction_mode == "page_image_ocr":
        return _extract_pdf_text_with_page_image_ocr(content, filename, app_config or {}, ollama_url)

    if extraction_mode == "qwen_vl":
        return _extract_pdf_text_with_vision_ocr(content, app_config or {}, ollama_url)

    if extraction_mode == "surya_docling":
        embedded_text = _extract_pdf_embedded_text(content)
        if len(embedded_text.strip()) >= PDF_FAST_TEXT_MIN_CHARS:
            return embedded_text

        try:
            advanced_text = _extract_pdf_text_with_surya_docling(content, filename)
            if advanced_text.strip():
                return advanced_text
        except Exception as exc:
            print(f"[Document Parser] Surya/Docling extraction failed, falling back to legacy PDF extraction: {exc}")

    return _extract_pdf_text_legacy(content)


def _pdf_extraction_mode(app_config: dict) -> str:
    configured_pdf_mode = str(app_config.get("pdf_extraction_mode") or "").strip()
    if configured_pdf_mode == "auto":
        configured_pdf_mode = ""
    if configured_pdf_mode in {"surya_docling", "legacy", "qwen_vl", "page_image_ocr"}:
        return configured_pdf_mode

    configured_engine = str(app_config.get("ocr_engine") or "").strip()
    if configured_engine in {"auto", "surya_docling"}:
        return "surya_docling"
    if configured_engine == "qwen_vl":
        return "qwen_vl"
    if configured_engine == "tesseract":
        return "legacy"
    return PDF_EXTRACTION_MODE


def _extract_pdf_text_with_page_image_ocr(content: bytes, filename: str, app_config: dict, ollama_url: str | None) -> str:
    page_count = _pdf_page_count(content)
    try:
        vision_text = _extract_pdf_page_images_with_vision_ocr(content, app_config, ollama_url)
        if _has_all_pdf_page_markers(vision_text, page_count):
            return vision_text
        if vision_text.strip():
            raise RuntimeError(f"Qwen-VL OCR returned incomplete page text: expected {page_count} pages")
    except Exception as exc:
        print(f"[Document Parser] Qwen-VL PDF page-image OCR failed, trying Surya OCR: {exc}")

    try:
        surya_text = _extract_pdf_with_surya(content, filename)
        if surya_text.strip():
            combined_surya_text = "OCR Result (Surya)\n" + surya_text.strip()
            if _has_all_pdf_page_markers(combined_surya_text, page_count):
                return combined_surya_text
            raise RuntimeError(f"Surya OCR returned incomplete page text: expected {page_count} pages")
    except Exception as exc:
        print(f"[Document Parser] Surya OCR fallback failed, trying Tesseract page-image OCR: {exc}")

    return _extract_pdf_page_images_with_tesseract(content)


def _pdf_page_count(content: bytes) -> int:
    with pymupdf.open(stream=content, filetype="pdf") as document:
        return document.page_count


def _has_all_pdf_page_markers(text: str, page_count: int) -> bool:
    if page_count <= 0:
        return bool(text.strip())
    return all(f"Page {page_number}" in text for page_number in range(1, page_count + 1))


def _extract_pdf_page_images_with_vision_ocr(content: bytes, app_config: dict, ollama_url: str | None) -> str:
    text_parts = []
    with pymupdf.open(stream=content, filetype="pdf") as document:
        for page_index in range(document.page_count):
            image_bytes = _render_pdf_page_to_jpg(document.load_page(page_index))
            text = _extract_image_text_with_vision_model(image_bytes, app_config, ollama_url).strip()
            if text:
                text_parts.append(f"Page {page_index + 1}\n{text}")
    return "\n\n".join(text_parts)


def _extract_pdf_page_images_with_tesseract(content: bytes) -> str:
    text_parts = []
    with pymupdf.open(stream=content, filetype="pdf") as document:
        for page_index in range(document.page_count):
            image_bytes = _render_pdf_page_to_jpg(document.load_page(page_index))
            text = _extract_image_text_with_tesseract(image_bytes).strip()
            text_parts.append(f"Page {page_index + 1}\n{text or '[No OCR text extracted]'}")
    return "\n\n".join(text_parts)


def _extract_pdf_text_with_vision_ocr(content: bytes, app_config: dict, ollama_url: str | None) -> str:
    reader = PdfReader(io.BytesIO(content))
    text_parts = []
    with pymupdf.open(stream=content, filetype="pdf") as ocr_document:
        for index, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if not text:
                image_bytes = _render_pdf_page_to_png(ocr_document.load_page(index - 1))
                try:
                    text = _extract_image_text_with_vision_model(image_bytes, app_config, ollama_url).strip()
                except Exception as exc:
                    print(f"[Document Parser] Vision PDF OCR failed on page {index}, falling back to Tesseract: {exc}")
                    text = _ocr_pdf_page(ocr_document.load_page(index - 1)).strip()
            if text:
                text_parts.append(f"Page {index}\n{text}")
    return "\n\n".join(text_parts)


def _render_pdf_page_to_png(page) -> bytes:
    pixmap = page.get_pixmap(dpi=PDF_OCR_DPI, alpha=False)
    return pixmap.tobytes("png")


def _render_pdf_page_to_jpg(page) -> bytes:
    pixmap = page.get_pixmap(dpi=PDF_OCR_DPI, alpha=False)
    with Image.open(io.BytesIO(pixmap.tobytes("png"))) as image:
        buffer = io.BytesIO()
        image.convert("RGB").save(buffer, format="JPEG", quality=95)
        return buffer.getvalue()


def _extract_pdf_embedded_text(content: bytes) -> str:
    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception:
        return ""

    text_parts = []
    for index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            text_parts.append(f"Page {index}\n{text}")
    return "\n\n".join(text_parts)


def _extract_pdf_text_with_surya_docling(content: bytes, filename: str) -> str:
    docling_text = _extract_pdf_with_docling(content, filename)
    if len(docling_text.strip()) >= PDF_SURYA_MIN_DOCLING_CHARS:
        return _combine_advanced_pdf_sections(docling_text, "")

    surya_text = _extract_pdf_with_surya(content, filename)
    return _combine_advanced_pdf_sections(docling_text, surya_text)


def _combine_advanced_pdf_sections(docling_text: str, surya_text: str) -> str:
    combined_sections = []

    if docling_text.strip():
        combined_sections.append("Document Parsing Result (Docling)\n" + docling_text.strip())
    if surya_text.strip():
        combined_sections.append("OCR Result (Surya)\n" + surya_text.strip())

    if not combined_sections:
        return ""

    combined_text = "\n\n".join(combined_sections)
    if len(combined_text) > PDF_ADVANCED_MAX_CHARS:
        return combined_text[:PDF_ADVANCED_MAX_CHARS].rstrip() + "\n\n[Advanced PDF extraction truncated.]"
    return combined_text


def _extract_pdf_with_docling(content: bytes, filename: str) -> str:
    try:
        from docling.document_converter import DocumentConverter
    except ImportError as exc:
        raise RuntimeError("Docling is not installed") from exc

    with tempfile.TemporaryDirectory() as temp_dir:
        pdf_path = Path(temp_dir) / _safe_temp_pdf_name(filename)
        pdf_path.write_bytes(content)
        converter = DocumentConverter()
        result = converter.convert(str(pdf_path))
        return result.document.export_to_markdown()


def _extract_pdf_with_surya(content: bytes, filename: str) -> str:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        pdf_path = temp_path / _safe_temp_pdf_name(filename)
        output_dir = temp_path / "surya_output"
        pdf_path.write_bytes(content)
        command = [
            "surya_ocr",
            str(pdf_path),
            "--results_dir",
            str(output_dir),
        ]
        max_pages = os.getenv("PDF_SURYA_MAX_PAGES", "").strip()
        if max_pages:
            command.extend(["--max", max_pages])
        start_page = os.getenv("PDF_SURYA_START_PAGE", "").strip()
        if start_page:
            command.extend(["--start_page", start_page])

        completed = _run_surya_ocr_command(command)
        if completed.returncode != 0 and _should_retry_surya_on_cpu(completed):
            print("[Document Parser] Surya GPU extraction failed, retrying Surya on CPU.")
            completed = _run_surya_ocr_command(command, force_cpu=True)
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "Surya OCR failed").strip())

        results_file = output_dir / "results.json"
        if not results_file.exists():
            candidate_files = list(output_dir.rglob("results.json"))
            if not candidate_files:
                return ""
            results_file = candidate_files[0]

        return _extract_text_from_surya_results(json.loads(results_file.read_text(encoding="utf-8")))


def _run_surya_ocr_command(command: list[str], force_cpu: bool = False):
    env = os.environ.copy()
    if force_cpu:
        env["TORCH_DEVICE"] = "cpu"
    return subprocess.run(command, capture_output=True, text=True, timeout=PDF_SURYA_TIMEOUT_SECONDS, check=False, env=env)


def _should_retry_surya_on_cpu(completed) -> bool:
    if os.getenv("PDF_SURYA_DISABLE_CPU_RETRY", "").lower() in {"1", "true", "yes"}:
        return False

    output = f"{completed.stderr or ''}\n{completed.stdout or ''}".lower()
    return any(
        marker in output
        for marker in (
            "cuda out of memory",
            "found no nvidia driver",
            "torch.cuda",
            "cuda error",
        )
    )


def _safe_temp_pdf_name(filename: str) -> str:
    stem = Path(filename or "document.pdf").stem or "document"
    safe_stem = "".join(char if char.isalnum() or char in {"-", "_"} else "_" for char in stem)[:80] or "document"
    return f"{safe_stem}.pdf"


def _extract_text_from_surya_results(payload) -> str:
    text_parts = []

    def visit(value):
        if isinstance(value, dict):
            for key, child in value.items():
                if key in {"text", "html", "markdown"} and isinstance(child, str) and child.strip():
                    text_parts.append(child.strip())
                else:
                    visit(child)
        elif isinstance(value, list):
            for child in value:
                visit(child)

    visit(payload)
    return "\n".join(dict.fromkeys(text_parts))


def _extract_pdf_text_legacy(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    text_parts = []
    with pymupdf.open(stream=content, filetype="pdf") as ocr_document:
        for index, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if not text:
                text = _ocr_pdf_page(ocr_document.load_page(index - 1)).strip()
            if text:
                text_parts.append(f"Page {index}\n{text}")
    return "\n\n".join(text_parts)


def _ocr_pdf_page(page) -> str:
    pixmap = page.get_pixmap(dpi=PDF_OCR_DPI, alpha=False)
    image = Image.open(io.BytesIO(pixmap.tobytes("png")))
    try:
        return pytesseract.image_to_string(image, lang=PDF_OCR_LANGUAGE)
    finally:
        image.close()


def _extract_image_text(content: bytes, app_config: dict | None = None, ollama_url: str | None = None) -> str:
    text, _metadata = _extract_image_text_with_metadata(content, app_config, ollama_url)
    return text


def _extract_image_text_with_metadata(content: bytes, app_config: dict | None = None, ollama_url: str | None = None) -> tuple[str, dict]:
    config = app_config or {}
    ocr_engine = str(config.get("ocr_engine") or "auto").strip()

    if ocr_engine in {"auto", "qwen_vl"}:
        try:
            vision_text = _extract_image_text_with_vision_model(content, config, ollama_url)
            if vision_text.strip():
                return vision_text, {
                    "ocr_engine_used": "qwen_vl",
                    "ocr_model": str(config.get("vision_ocr_model") or "").strip(),
                }
        except Exception as exc:
            print(f"[Document Parser] Vision OCR failed, falling back to Tesseract: {exc}")
            return _extract_image_text_with_tesseract(content), {
                "ocr_engine_used": "tesseract",
                "ocr_fallback_from": "qwen_vl",
                "ocr_fallback_reason": str(exc),
            }

    return _extract_image_text_with_tesseract(content), {"ocr_engine_used": "tesseract"}


def _extract_image_text_with_vision_model(content: bytes, app_config: dict, ollama_url: str | None) -> str:
    if not ollama_url:
        raise RuntimeError("Ollama URL is not configured")

    model = str(app_config.get("vision_ocr_model") or "").strip()
    if not model:
        raise RuntimeError("vision_ocr_model is not configured")

    timeout = int(app_config.get("vision_ocr_timeout_seconds", 120) or 120)
    prompt = str(app_config.get("vision_ocr_prompt") or DEFAULT_VISION_OCR_PROMPT).strip()
    payload = {
        "model": model,
        "stream": False,
        "messages": [
            {
                "role": "user",
                "content": prompt,
                "images": [base64.b64encode(content).decode("ascii")],
            }
        ],
        "options": {
            "temperature": 0,
        },
    }
    apply_gpu_defaults(payload)
    response = requests.post(f"{ollama_url}/api/chat", json=payload, timeout=timeout)
    response.raise_for_status()
    return str(response.json().get("message", {}).get("content", "")).strip()


def _extract_image_text_with_tesseract(content: bytes) -> str:
    with Image.open(io.BytesIO(content)) as image:
        ocr_image = image.convert("RGB") if image.mode not in {"RGB", "L"} else image
        try:
            return pytesseract.image_to_string(ocr_image, lang=IMAGE_OCR_LANGUAGE)
        finally:
            if ocr_image is not image:
                ocr_image.close()


def _extract_docx_text(content: bytes) -> str:
    docx_file = io.BytesIO(content)
    doc = Document(docx_file)
    return "\n".join(para.text for para in doc.paragraphs if para.text)


def _extract_xlsx_text(content: bytes) -> str:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    return _render_workbook_sheets(workbook.sheetnames, lambda name: workbook[name].iter_rows(values_only=True))


def _extract_xls_text(content: bytes) -> str:
    workbook = xlrd.open_workbook(file_contents=content)

    def iter_rows(sheet_name: str):
        sheet = workbook.sheet_by_name(sheet_name)
        for row_index in range(sheet.nrows):
            yield sheet.row_values(row_index)

    return _render_workbook_sheets(workbook.sheet_names(), iter_rows)


def _render_workbook_sheets(sheet_names, row_provider) -> str:
    parts = []
    for sheet_name in sheet_names:
        parts.append(f"Sheet: {sheet_name}")
        for row in row_provider(sheet_name):
            row_text = ", ".join(str(value) if value is not None else "" for value in row).strip(", ")
            if row_text:
                parts.append(row_text)
        parts.append("")
    return "\n".join(parts).strip()


def _truncate_document_text(file_text: str) -> str:
    if len(file_text) <= MAX_DOCUMENT_PROMPT_CHARS:
        return file_text

    return (
        file_text[:MAX_DOCUMENT_PROMPT_CHARS].rstrip() +
        f"\n\n[Content truncated to the first {MAX_DOCUMENT_PROMPT_CHARS} characters to fit the model context window.]"
    )
